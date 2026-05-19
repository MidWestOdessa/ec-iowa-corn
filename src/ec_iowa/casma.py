"""Crop-CASMA client: NASA SMAP weekly soil moisture via CSISS @ GMU.

For one (county, ISO-week, depth) combination there are three places the
data might live, in increasing cost:

  1. Local file cache         (cache/casma/{layer}_{fips}.csv)
  2. Server cache CSV         (smap_cache/byFips/...)
  3. Server WPS Execute call  (computes + caches, returns CSV URL)

`fetch_county_week` walks these in order and returns the raw pixel-count
dict (category code -> pixel count). Categories: 0=no-data, 1=Very Short,
2=Short, 3=Adequate, 4=Surplus.

The WPS DataInputs format was reverse-engineered against the live PyWPS
4.4.2 server: `layer=...;fips=...;minValue=1.0;maxValue=4.0;step=1.0`.
"""
from __future__ import annotations

import time
import urllib.parse as up
import xml.etree.ElementTree as ET
from collections.abc import Mapping
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING, Literal

import requests
from openpyxl.comments import Comment

from ec_iowa import config

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook

Depth = Literal["TOP", "SUB"]
PixelCounts = dict[int, int]
Percentages = tuple[float, float, float, float]  # (VS, S, A, Su)


@dataclass(frozen=True)
class DistrictRollup:
    """Corn-acre-weighted district percentages for one depth, plus coverage metadata."""

    depth: Depth
    pcts: Percentages
    counties_included: list[str] = field(default_factory=list)
    acres_covered: int = 0
    acres_total: int = 0

    @property
    def coverage_fraction(self) -> float:
        return self.acres_covered / self.acres_total if self.acres_total else 0.0

    def coverage_note(self) -> str:
        n = len(self.counties_included)
        pct = self.coverage_fraction * 100
        return f"{n}/9 counties, {pct:.0f}% of corn acres"

CASMA_CACHE_DIR = config.CACHE_DIR / "casma"
WPS_NS = {"wps": "http://www.opengis.net/wps/1.0.0", "ows": "http://www.opengis.net/ows/1.1"}


# ---- URL + name helpers -------------------------------------------------

def iso_week_dates(year: int, iso_week: int) -> tuple[date, date]:
    """Return (Monday, Sunday) of the given ISO week."""
    monday = date.fromisocalendar(year, iso_week, 1)
    return monday, monday + timedelta(days=6)


def layer_name(year: int, iso_week: int, depth: Depth, monday: date, sunday: date) -> str:
    """e.g. SMAP-9KM-CATEGORY-WEEKLY-TOP_2026_14_2026.03.30_2026.04.05_AVERAGE"""
    fmt = lambda d: f"{d.year:04d}.{d.month:02d}.{d.day:02d}"
    return (
        f"SMAP-9KM-CATEGORY-WEEKLY-{depth}"
        f"_{year}_{iso_week:02d}"
        f"_{fmt(monday)}_{fmt(sunday)}"
        f"_AVERAGE"
    )


def cache_csv_url(layer: str, fips: str) -> str:
    return f"{config.CASMA_CACHE_BASE}{layer}_{fips}/{layer}_{fips}_1.0_4.0_1.0.csv"


def wps_execute_url(layer: str, fips: str) -> str:
    data_inputs = (
        f"layer={layer};fips={fips};minValue=1.0;maxValue=4.0;step=1.0"
    )
    params = {
        "service": "WPS",
        "version": "1.0.0",
        "request": "Execute",
        "identifier": "GetStatByFips",
        "DataInputs": data_inputs,
    }
    # Preserve `;` and `=` and `:` and `.` literally; the PyWPS server
    # parses DataInputs with semicolons as separators.
    safe = ";=,@.-_:/"
    qs = "&".join(f"{k}={up.quote(str(v), safe=safe)}" for k, v in params.items())
    return f"{config.CASMA_WPS_BASE}?{qs}"


# ---- CSV parsing --------------------------------------------------------

def parse_csv(text: str) -> PixelCounts:
    """`category,pixels` -> {0: ..., 1: ..., ...}. Tolerates blank lines."""
    out: PixelCounts = {}
    for line_no, line in enumerate(text.splitlines()):
        if line_no == 0 or not line.strip():
            continue
        parts = line.split(",")
        if len(parts) != 2:
            raise ValueError(f"Bad CSV line {line_no}: {line!r}")
        out[int(parts[0])] = int(parts[1])
    return out


def percentages(pixels: PixelCounts) -> Percentages | None:
    """(VS, S, A, Su) % of valid pixels (cats 1-4). None if all-no-data."""
    total = sum(pixels.get(c, 0) for c in (1, 2, 3, 4))
    if total == 0:
        return None
    return tuple(round(100 * pixels.get(c, 0) / total, 2) for c in (1, 2, 3, 4))  # type: ignore[return-value]


# ---- HTTP with retries --------------------------------------------------

class CasmaError(RuntimeError):
    """Raised when CASMA can't be fetched after retries."""


class CasmaDataNotAvailable(CasmaError):
    """The WPS server returned ProcessFailed — data hasn't been processed yet.

    SMAP weekly aggregations have ~1-2 week processing latency at CSISS.
    Recent weeks frequently raise this; retry later.
    """


def _get_with_retry(
    url: str,
    *,
    timeout_s: int = 30,
    retries: int = 3,
    backoff_s: float = 2.0,
    accept_404: bool = False,
) -> requests.Response:
    """GET with exponential backoff. If accept_404, returns 404 responses
    without retrying or raising — caller decides what to do.

    Also accepts HTTP 400 responses whose body contains an XML
    ExceptionReport — these come from PyWPS when the requested data
    hasn't been processed yet. The caller (which expects to parse a WPS
    response anyway) gets to detect that and raise CasmaDataNotAvailable.
    """
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=timeout_s)
            if r.status_code == 200:
                return r
            if accept_404 and r.status_code == 404:
                return r
            # HTTP 400 with a PyWPS exception report → let caller parse it.
            if r.status_code == 400 and "ExceptionReport" in r.text:
                return r
            # Anything else: retry on 5xx, raise on 4xx
            if 500 <= r.status_code < 600:
                last_exc = CasmaError(f"HTTP {r.status_code} from {url}")
            else:
                raise CasmaError(f"HTTP {r.status_code} from {url}: {r.text[:200]}")
        except requests.RequestException as exc:
            last_exc = exc
        if attempt < retries - 1:
            time.sleep(backoff_s * (2 ** attempt))
    assert last_exc is not None
    raise CasmaError(f"failed after {retries} attempts: {last_exc!r}") from last_exc


def _wps_extract_output_url(xml_text: str) -> str:
    """Pull the outputUrl LiteralData out of a WPS Execute response.

    Raises CasmaDataNotAvailable for two known error shapes:
      1. <ows:ExceptionReport> at the document root (HTTP 400 path,
         observed at CSISS as of ~mid-May 2026)
      2. <wps:ProcessFailed> inside an ExecuteResponse (HTTP 200 path,
         the original observed behavior)
    """
    root = ET.fromstring(xml_text)
    # Case 1: top-level ExceptionReport
    if root.tag == f"{{{WPS_NS['ows']}}}ExceptionReport":
        ex_text = root.find(f".//{{{WPS_NS['ows']}}}ExceptionText")
        detail = ex_text.text.strip() if (ex_text is not None and ex_text.text) else "WPS ExceptionReport"
        raise CasmaDataNotAvailable(detail)
    # Case 2: ProcessFailed inside ExecuteResponse
    if root.find(f".//{{{WPS_NS['wps']}}}ProcessFailed") is not None:
        ex_text = root.find(f".//{{{WPS_NS['ows']}}}ExceptionText")
        detail = ex_text.text.strip() if (ex_text is not None and ex_text.text) else "ProcessFailed"
        raise CasmaDataNotAvailable(detail)
    for output in root.iter(f"{{{WPS_NS['wps']}}}Output"):
        ident = output.find(f"{{{WPS_NS['ows']}}}Identifier")
        if ident is not None and ident.text == "outputUrl":
            data = output.find(f"{{{WPS_NS['wps']}}}Data/{{{WPS_NS['wps']}}}LiteralData")
            if data is not None and data.text:
                return data.text.strip()
    raise CasmaError("WPS response missing outputUrl LiteralData")


# ---- Public API ---------------------------------------------------------

def fetch_county_week(
    fips: str,
    year: int,
    iso_week: int,
    depth: Depth,
    *,
    use_local_cache: bool = True,
    timeout_s: int = 30,
) -> PixelCounts:
    """Fetch SMAP weekly soil moisture pixel counts for one (county, week, depth).

    Walks: local file cache -> server CSV cache -> server WPS trigger -> server CSV.
    """
    monday, sunday = iso_week_dates(year, iso_week)
    layer = layer_name(year, iso_week, depth, monday, sunday)

    local_path = CASMA_CACHE_DIR / f"{layer}_{fips}.csv"
    if use_local_cache and local_path.exists():
        return parse_csv(local_path.read_text(encoding="utf-8"))

    # 1) Try the cached CSV directly (someone else might have triggered it).
    csv_url = cache_csv_url(layer, fips)
    r = _get_with_retry(csv_url, timeout_s=timeout_s, accept_404=True)
    if r.status_code == 200:
        return _persist_and_parse(local_path, r.text)

    # 2) WPS trigger; the server computes and returns the (now-existing) CSV URL.
    wps_url = wps_execute_url(layer, fips)
    wps_resp = _get_with_retry(wps_url, timeout_s=timeout_s)
    output_url = _wps_extract_output_url(wps_resp.text)

    # 3) Fetch the freshly-cached CSV.
    csv_resp = _get_with_retry(output_url, timeout_s=timeout_s)
    return _persist_and_parse(local_path, csv_resp.text)


def _persist_and_parse(local_path: Path, csv_text: str) -> PixelCounts:
    local_path.parent.mkdir(parents=True, exist_ok=True)
    local_path.write_text(csv_text, encoding="utf-8")
    return parse_csv(csv_text)


def fetch_district_week(
    year: int,
    iso_week: int,
    *,
    depths: tuple[Depth, ...] = ("TOP", "SUB"),
    use_local_cache: bool = True,
    timeout_s: int = 30,
    on_failure: Literal["raise", "skip"] = "raise",
) -> dict[tuple[str, Depth], PixelCounts]:
    """Fetch all EC-Iowa counties for the given week + depth(s), sequentially.

    Sequential because CASMA WPS calls can be expensive server-side and we'd
    rather be polite than fast. ~1.3 s per HTTP miss; local cache is ~25 ms.

    on_failure='skip' lets weekly-update tolerate single-county hiccups.
    """
    out: dict[tuple[str, Depth], PixelCounts] = {}
    failures: list[tuple[str, Depth, Exception]] = []
    for fips in config.EC_IOWA_COUNTIES:
        for depth in depths:
            try:
                out[(fips, depth)] = fetch_county_week(
                    fips, year, iso_week, depth,
                    use_local_cache=use_local_cache, timeout_s=timeout_s,
                )
            except Exception as exc:
                if on_failure == "raise":
                    raise
                failures.append((fips, depth, exc))
    if failures and on_failure == "skip":
        # Surface the failures via a side-channel attribute on the dict
        out["_failures"] = failures  # type: ignore[assignment]
    return out


def compute_district_rollup(
    per_county: Mapping[tuple[str, Depth], PixelCounts],
    depth: Depth,
    *,
    county_acres: Mapping[str, int] | None = None,
) -> DistrictRollup | None:
    """Corn-acre-weighted average of per-county percentages for one depth.

    Skips counties whose pixel data is all no-data (no valid pixels in
    categories 1-4). Returns None if every county is no-data.
    """
    if county_acres is None:
        county_acres = {fips: acres for fips, (_, acres) in config.EC_IOWA_COUNTIES.items()}

    weighted = [0.0, 0.0, 0.0, 0.0]
    total_weight = 0
    used: list[str] = []

    for fips, acres in county_acres.items():
        pixels = per_county.get((fips, depth))
        if pixels is None:
            continue
        pcts = percentages(pixels)
        if pcts is None:
            continue
        for i, p in enumerate(pcts):
            weighted[i] += p * acres
        total_weight += acres
        used.append(fips)

    if total_weight == 0:
        return None

    rolled: Percentages = tuple(round(w / total_weight, 2) for w in weighted)  # type: ignore[assignment]
    acres_total = sum(county_acres.values())
    return DistrictRollup(
        depth=depth,
        pcts=rolled,
        counties_included=used,
        acres_covered=total_weight,
        acres_total=acres_total,
    )


# ---- Workbook write -----------------------------------------------------

# Crop-CASMA archive layout (verified 2026-05-01 against v5 workbook):
#   Header at row 95.  Data rows 96-131 (Mar 30 2026 -> Nov 30 2026, weekly).
#   Col A: Monday date.  Col B: ISO week #.
#   Cols C-F: Top VS / S / A / Su.   Cols G-J: Sub VS / S / A / Su.
_CASMA_ARCHIVE_FIRST_ROW = 96
_CASMA_ARCHIVE_LAST_ROW = 131
_CASMA_TOP_COL_BASE = 3   # Top VS at column C
_CASMA_SUB_COL_BASE = 7   # Sub VS at column G


def _find_archive_row(ws, monday: date) -> int:
    """Locate the archive row whose date column matches `monday`."""
    for row in range(_CASMA_ARCHIVE_FIRST_ROW, _CASMA_ARCHIVE_LAST_ROW + 1):
        cell_date = ws.cell(row, 1).value
        if isinstance(cell_date, datetime):
            cell_date = cell_date.date()
        if cell_date == monday:
            return row
    raise ValueError(
        f"No Crop-CASMA archive row matches Monday {monday} "
        f"(searched rows {_CASMA_ARCHIVE_FIRST_ROW}-{_CASMA_ARCHIVE_LAST_ROW})"
    )


def casma_to_nass_substress(casma_substress: float) -> float:
    """Translate CASMA subsoil VS+S percentage to its NASS-equivalent.

    CASMA reads systematically lower than NASS (satellite vs human raters);
    yield-model coefficients were trained on NASS values, so 2026+ CASMA
    inputs need this calibration before being fed to the yield model.

    See config.CASMA_NASS_SUBSTRESS_CALIBRATION for fit details.
    """
    cal = config.CASMA_NASS_SUBSTRESS_CALIBRATION
    return float(cal["intercept"]) + float(cal["slope"]) * float(casma_substress)


def write_to_archive(
    wb: "Workbook",
    year: int,
    iso_week: int,
    top_rollup: DistrictRollup,
    sub_rollup: DistrictRollup,
) -> int:
    """Write the 8 percentages to the Crop-CASMA archive row matching the Monday.

    Adds a comment on column A if either rollup has partial coverage.
    Returns the row number written (handy for tests / logging).
    """
    if top_rollup.depth != "TOP" or sub_rollup.depth != "SUB":
        raise ValueError("top_rollup must be depth=TOP and sub_rollup must be depth=SUB")
    ws = wb[config.SHEET_CASMA]
    monday, _ = iso_week_dates(year, iso_week)
    target_row = _find_archive_row(ws, monday)

    ws.cell(target_row, 2, iso_week)
    for i, pct in enumerate(top_rollup.pcts):
        ws.cell(target_row, _CASMA_TOP_COL_BASE + i, pct)
    for i, pct in enumerate(sub_rollup.pcts):
        ws.cell(target_row, _CASMA_SUB_COL_BASE + i, pct)

    if top_rollup.coverage_fraction < 1.0 or sub_rollup.coverage_fraction < 1.0:
        msg = f"TOP: {top_rollup.coverage_note()}\nSUB: {sub_rollup.coverage_note()}"
        ws.cell(target_row, 1).comment = Comment(msg, "ec_iowa")
    else:
        # Full coverage — clear any stale partial-coverage comment.
        ws.cell(target_row, 1).comment = None

    return target_row
