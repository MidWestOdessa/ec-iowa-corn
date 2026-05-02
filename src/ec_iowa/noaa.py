"""NOAA CDO daily temps for Cedar Rapids (USW00014990) + GDD50 computation.

Daily GDD50 = max(0, (max(min(TMAX, 86), 50) + max(min(TMIN, 86), 50)) / 2 - 50)
Cumulative from May 1.  See handoff §6.1.2.

Public API:
  fetch_daily_temps(station_id, start, end)            -> {date: {'TMAX': F, 'TMIN': F}}
  compute_gdd50_daily(tmax_f, tmin_f)                  -> float
  cumulative_gdd(daily_temps, accum_start, last_day)   -> {date: cumulative GDD50}
  write_to_workbook(wb, year, cumulative_gdd, accum_start) -> int (cells written)
"""
from __future__ import annotations

import os
from collections.abc import Mapping
from datetime import date, datetime, timedelta
from typing import TYPE_CHECKING

import requests
from dotenv import load_dotenv

from ec_iowa import config

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook


CDO_DATA_URL = "https://www.ncdc.noaa.gov/cdo-web/api/v2/data"


class NoaaError(RuntimeError):
    pass


def _get_token() -> str:
    load_dotenv()
    token = os.environ.get("NOAA_TOKEN")
    if not token:
        raise NoaaError("NOAA_TOKEN not set. Add to .env (template in .env.example).")
    return token


def fetch_daily_temps(
    station_id: str,
    start: date,
    end: date,
    *,
    token: str | None = None,
    timeout_s: int = 30,
) -> dict[date, dict[str, float]]:
    """Daily TMAX/TMIN (Fahrenheit) for the date range, inclusive.

    NOAA CDO limits a single request to 1 year. Caller must chunk for longer.
    Days with missing TMAX/TMIN simply don't appear in the result.
    """
    if (end - start).days > 365:
        raise NoaaError(
            f"date range {(end - start).days} days exceeds CDO 1-year limit"
        )
    if token is None:
        token = _get_token()
    sid = station_id if station_id.startswith("GHCND:") else f"GHCND:{station_id}"
    base_params = {
        "datasetid": "GHCND",
        "stationid": sid,
        "startdate": start.isoformat(),
        "enddate": end.isoformat(),
        "datatypeid": ["TMAX", "TMIN"],
        "units": "standard",   # Fahrenheit
        "limit": 1000,
    }
    out: dict[date, dict[str, float]] = {}
    offset = 1
    while True:
        params = {**base_params, "offset": offset}
        r = requests.get(
            CDO_DATA_URL, params=params, headers={"token": token}, timeout=timeout_s
        )
        if r.status_code != 200:
            raise NoaaError(f"HTTP {r.status_code}: {r.text[:200]}")
        data = r.json()
        results = data.get("results", [])
        for rec in results:
            d = datetime.fromisoformat(rec["date"]).date()
            out.setdefault(d, {})[rec["datatype"]] = float(rec["value"])
        meta = data.get("metadata", {}).get("resultset", {})
        total = meta.get("count", len(results))
        offset += len(results)
        if not results or offset > total:
            break
    return out


def compute_gdd50_daily(
    tmax_f: float, tmin_f: float, *, base_f: int = config.GDD_BASE_F, cap_f: int = config.GDD_CAP_HIGH_F
) -> float:
    """Daily GDD50: cap each temp at [base, cap], average, subtract base, floor at 0."""
    tmax_c = max(min(tmax_f, cap_f), base_f)
    tmin_c = max(min(tmin_f, cap_f), base_f)
    return max(0.0, (tmax_c + tmin_c) / 2 - base_f)


def cumulative_gdd(
    daily_temps: Mapping[date, Mapping[str, float]],
    accum_start: date,
    last_day: date,
) -> dict[date, float]:
    """Running cumulative GDD50 for every day in [accum_start, last_day], rounded to 0.1."""
    out: dict[date, float] = {}
    cum = 0.0
    d = accum_start
    while d <= last_day:
        if d in daily_temps:
            t = daily_temps[d]
            tmax, tmin = t.get("TMAX"), t.get("TMIN")
            if tmax is not None and tmin is not None:
                cum += compute_gdd50_daily(tmax, tmin)
        out[d] = round(cum, 1)
        d += timedelta(days=1)
    return out


def write_to_workbook(
    wb: "Workbook",
    year: int,
    cumulative: Mapping[date, float],
    *,
    accum_start: date,
    week_end_offset_days: int = 6,
    last_col: int = 37,  # AK; matches handoff §2 (155x37 dims)
) -> int:
    """Write cumulative GDD to the year's GDD row of the Crop Progress block.

    Each Monday column M reports cumulative GDD AS OF the Sunday ending that
    ISO week (= M + week_end_offset_days = M + 6 days), matching the NASS
    crop-progress weekly-report convention used elsewhere in the workbook.

    For each populated Monday in the dates row:
      - If Sunday < accum_start: writes 0
      - Else if Sunday in `cumulative`: writes that value
      - Else: skips (likely a future Sunday we don't have data for yet)

    Returns the number of cells written.
    """
    ws = wb[config.SHEET_CROP_PROGRESS]
    block = config.CROP_PROGRESS_YEAR_BLOCKS[year]
    dates_row, gdd_row = block["dates"], block["gdd"]

    written = 0
    for col in range(2, last_col + 1):
        cell_date = ws.cell(dates_row, col).value
        if cell_date is None:
            continue
        if isinstance(cell_date, datetime):
            cell_date = cell_date.date()
        if not isinstance(cell_date, date):
            continue
        target = cell_date + timedelta(days=week_end_offset_days)
        if target < accum_start:
            ws.cell(gdd_row, col).value = 0
            written += 1
        elif target in cumulative:
            ws.cell(gdd_row, col).value = cumulative[target]
            written += 1
    return written
