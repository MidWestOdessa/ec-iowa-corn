"""Workbook safety helpers.

Every script that writes to the canonical workbook must: refuse to run while
the owner has it open in Excel, take a dated backup first, and locate columns
by date rather than hard-coded letters. Those three things were re-implemented
in ~80 ad-hoc scripts; they live here now.

Two openpyxl behaviours to keep in mind (MODEL_HANDOFF.md P2, P3):
  * `data_only=True` returns Excel's LAST CACHED value for a formula cell. If
    openpyxl wrote the upstream data and the file hasn't been reopened in
    Excel since, dependants still read None. Read source-of-truth literals.
  * Saving drops conditional formatting extensions (~7% size loss). Harmless,
    but back up before the first save of a session.

Public API:
    ensure_writable()                 -> None      (raises WorkbookLocked)
    backup(tag)                       -> Path
    load(data_only=False)             -> Workbook
    open_for_write(tag)               -> (Workbook, backup_path)
    column_for_monday(ws, dates_row, monday) -> int
    set_with_note(ws, row, col, value, note) -> None
"""
from __future__ import annotations

import shutil
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.comments import Comment

from ec_iowa import config

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook


class WorkbookLocked(RuntimeError):
    """The workbook is open in Excel and cannot be read or written."""


class WorkbookMissing(FileNotFoundError):
    """The canonical workbook is not present in this environment.

    Expected in a fresh clone or a cloud sandbox — the workbook is the
    deliverable and is not committed. Point EC_IOWA_WORKBOOK at a copy.
    """


def _missing(path: Path) -> WorkbookMissing:
    return WorkbookMissing(
        f"Workbook not found at:\n  {path}\n"
        "It is not in the repo (it lives in the owner's OneDrive). Set "
        "EC_IOWA_WORKBOOK to a local copy, or see .env.example."
    )


def ensure_writable(path: Path | None = None) -> None:
    """Raise if the workbook is absent, or held open by Excel."""
    path = path or config.WORKBOOK_PATH
    if not path.is_file():
        raise _missing(path)
    try:
        path.open("ab").close()
    except PermissionError as exc:
        raise WorkbookLocked(
            f"{path.name} is open in Excel — close it and retry."
        ) from exc


def backup(tag: str, path: Path | None = None) -> Path:
    """Copy the workbook to a dated sibling. Returns the backup path."""
    path = path or config.WORKBOOK_PATH
    stamp = date.today().isoformat()
    dest = path.with_name(f"{path.stem}.backup-{tag}-{stamp}.xlsx")
    shutil.copy2(path, dest)
    return dest


def load(data_only: bool = False, path: Path | None = None) -> "Workbook":
    """Load the workbook, suppressing openpyxl's extension-loss warnings.

    Excel holds an exclusive lock, so an open workbook blocks reads as well as
    writes. Translate that into WorkbookLocked rather than leaking a raw
    PermissionError with an unreadable path.
    """
    path = path or config.WORKBOOK_PATH
    if not path.is_file():
        raise _missing(path)
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            return openpyxl.load_workbook(path, data_only=data_only)
    except PermissionError as exc:
        raise WorkbookLocked(
            f"{path.name} is open in Excel — close it and retry."
        ) from exc


def open_for_write(tag: str, path: Path | None = None) -> tuple["Workbook", Path]:
    """Pre-flight, back up, and open for writing. The standard opening move."""
    path = path or config.WORKBOOK_PATH
    ensure_writable(path)
    bak = backup(tag, path)
    return load(data_only=False, path=path), bak


def column_for_monday(ws, dates_row: int, monday: date,
                      last_col: int = 37) -> int:
    """Column index whose dates-row cell equals `monday`.

    Locating columns by date avoids the hard-coded-letter drift that follows a
    column insert (MODEL_HANDOFF.md P1).
    """
    for col in range(2, last_col + 1):
        v = ws.cell(dates_row, col).value
        if isinstance(v, datetime):
            v = v.date()
        if v == monday:
            return col
    raise ValueError(f"no column in row {dates_row} matches Monday {monday}")


def set_with_note(ws, row: int, col: int, value, note: str,
                  author: str = "ec_iowa") -> None:
    """Write a value and attach a provenance comment.

    Manual and derived entries should always carry their origin, so a later
    reader can tell a fetched number from a judgement call.
    """
    cell = ws.cell(row, col)
    cell.value = value
    cell.comment = Comment(note, author)
