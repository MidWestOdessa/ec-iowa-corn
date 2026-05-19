"""Generate web/data.json from the canonical workbook + config.

Run weekly (or after a weekly-update) to refresh what the dashboard shows.
Usage:  uv run python -m web.snapshot
"""
from __future__ import annotations
import json, sys, warnings
from datetime import date, datetime
from pathlib import Path
from statistics import median

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
from ec_iowa import config

OUT = Path(__file__).parent / "data.json"


def _date_safe(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def main() -> None:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(config.WORKBOOK_PATH, data_only=True)
    ws_y = wb[config.SHEET_YIELD_MODEL]
    ws_cp = wb[config.SHEET_CROP_PROGRESS]
    ws_ca = wb[config.SHEET_CASMA]

    # ---- 1. Yield model history (training table rows 30-45) ----
    history = []
    for r in range(30, 46):
        yr = ws_y.cell(r, 1).value
        if not isinstance(yr, int):
            continue
        actual_raw = ws_y.cell(r, 2).value
        substr = ws_y.cell(r, 3).value
        ge = ws_y.cell(r, 4).value
        gdd = ws_y.cell(r, 5).value
        predicted = ws_y.cell(r, 7).value
        residual = ws_y.cell(r, 8).value
        note = ws_y.cell(r, 9).value
        actual = float(actual_raw) if isinstance(actual_raw, (int, float)) else None
        excluded = yr in config.YIELD_MODEL["training_excluded_years"]
        history.append({
            "year": yr,
            "actual": actual,
            "predicted": float(predicted) if isinstance(predicted, (int, float)) else None,
            "residual": float(residual) if isinstance(residual, (int, float)) else None,
            "substress_jul": float(substr) if isinstance(substr, (int, float)) else None,
            "excluded": excluded,
            "note": note,
        })

    # ---- 2. Indicative 2026 forecast ----
    # We don't have peak-July SubStress for 2026 yet. Use median of training years
    # as a placeholder, with uncertainty showing the trader the spread.
    train_substr = [
        h["substress_jul"] for h in history
        if h["substress_jul"] is not None and not h["excluded"] and h["year"] >= 2010 and h["year"] <= 2024
    ]
    median_substr = float(median(train_substr)) if train_substr else 20.0
    ym = config.YIELD_MODEL
    forecast_year = 2026
    indicative = (
        float(ym["intercept"])
        + float(ym["year"]) * forecast_year
        + float(ym["substress_jul"]) * median_substr
    )
    loocv_mae = float(ym["loocv_mae"])
    # ~95% band ≈ ±2σ; use 2*LOOCV_MAE as a practical envelope (conservative)
    forecast = {
        "year": forecast_year,
        "indicative_bu_ac": round(indicative, 1),
        "band_low_95":  round(indicative - 2 * loocv_mae, 1),
        "band_high_95": round(indicative + 2 * loocv_mae, 1),
        "loocv_mae":    round(loocv_mae, 2),
        "method":       "Year + SubStress_Jul (NASS-equivalent scale)",
        "inputs": {
            "year": forecast_year,
            "substress_jul_used": round(median_substr, 1),
            "substress_jul_source": (
                f"Median of 2010-2024 training SubStress_Jul ({len(train_substr)} years). "
                "Real 2026 peak-July value not yet observed — forecast firms up late July."
            ),
        },
        "calibration_note": (
            "When peak-July 2026 CASMA data lands, apply "
            "casma.casma_to_nass_substress() before plugging in here."
        ),
        "fit_stats": {
            "r_squared":  ym["r_squared"],
            "mae":        ym["mae"],
            "loocv_mae":  ym["loocv_mae"],
            "training_years": "2010-2024 excluding 2020, 2025",
            "excluded": ym["training_excluded_years"],
        },
    }

    # ---- 3. Current conditions snapshot ----
    # Latest CASMA archive row that's not empty
    latest_casma = None
    for r in range(131, 95, -1):
        if any(ws_ca.cell(r, c).value is not None for c in (3, 4, 5, 6, 7, 8, 9, 10)):
            latest_casma = {
                "monday":   _date_safe(ws_ca.cell(r, 1).value),
                "iso_week": ws_ca.cell(r, 2).value,
                "top": {
                    "vs": ws_ca.cell(r, 3).value, "s": ws_ca.cell(r, 4).value,
                    "a":  ws_ca.cell(r, 5).value, "su": ws_ca.cell(r, 6).value,
                },
                "sub": {
                    "vs": ws_ca.cell(r, 7).value, "s": ws_ca.cell(r, 8).value,
                    "a":  ws_ca.cell(r, 9).value, "su": ws_ca.cell(r, 10).value,
                },
            }
            break

    # 2026 GDD progression
    gdd_series = []
    dates_row = config.CROP_PROGRESS_YEAR_BLOCKS[2026]["dates"]
    gdd_row = config.CROP_PROGRESS_YEAR_BLOCKS[2026]["gdd"]
    for c in range(2, 38):
        d = ws_cp.cell(dates_row, c).value
        if d is None:
            continue
        g = ws_cp.cell(gdd_row, c).value
        if isinstance(g, (int, float)):
            gdd_series.append({"monday": _date_safe(d), "gdd": float(g)})

    # Crop progress (most recent non-zero values per stage)
    stages = ["planted", "emerged", "silking", "doughing", "dented", "corn_mature", "corn_harvested"]
    crop_progress = {}
    for stage in stages:
        row = dates_row + config.DATA_ROW_OFFSETS[stage]
        latest_val = latest_date = latest_wk = None
        for c in range(2, 38):
            d = ws_cp.cell(dates_row, c).value
            v = ws_cp.cell(row, c).value
            if isinstance(v, (int, float)) and v > 0 and d is not None:
                latest_val = float(v)
                latest_date = _date_safe(d)
                # compute ISO week from date
                if isinstance(d, datetime):
                    d = d.date()
                latest_wk = d.isocalendar()[1] if hasattr(d, "isocalendar") else None
        crop_progress[stage] = {
            "latest_pct": latest_val,
            "latest_monday": latest_date,
            "latest_iso_week": latest_wk,
        }

    conditions = {
        "latest_casma": latest_casma,
        "gdd_series": gdd_series,
        "crop_progress": crop_progress,
    }

    # ---- 4. Last 5 years weekly trajectories + 2026 ----
    # For 2021-2025: read soil moisture from Crop Progress (hardcoded literals).
    # For 2026: the Crop Progress soil rows are FORMULAS pulling from the
    # Crop-CASMA archive, and openpyxl-with-data_only returns whatever Excel
    # last cached (stale None if the user hasn't reopened in Excel after our
    # CASMA writes). So for 2026, bypass the formulas and read soil moisture
    # directly from the Crop-CASMA archive rows.
    comparison_years = [2021, 2022, 2023, 2024, 2025, 2026]
    history_weekly: dict[str, list[dict]] = {}
    stage_keys = ["planted", "emerged", "silking", "doughing", "dented", "corn_mature", "corn_harvested"]

    # Build a Monday-date -> Crop-CASMA archive row index for 2026.
    casma_row_by_monday: dict[date, int] = {}
    for r in range(96, 132):
        d_ca = ws_ca.cell(r, 1).value
        if isinstance(d_ca, datetime):
            d_ca = d_ca.date()
        if isinstance(d_ca, date):
            casma_row_by_monday[d_ca] = r

    def _safe_float(v):
        return float(v) if isinstance(v, (int, float)) else None

    for yr in comparison_years:
        if yr not in config.CROP_PROGRESS_YEAR_BLOCKS:
            continue
        blk = config.CROP_PROGRESS_YEAR_BLOCKS[yr]
        dr, gr = blk["dates"], blk["gdd"]
        weeks_out = []
        for c in range(2, 38):
            d = ws_cp.cell(dr, c).value
            if d is None:
                continue
            if isinstance(d, datetime):
                d = d.date()
            if not isinstance(d, date):
                continue
            iso_week = d.isocalendar()[1]
            entry = {
                "monday": d.isoformat(),
                "iso_week": iso_week,
                "gdd": ws_cp.cell(gr, c).value if isinstance(ws_cp.cell(gr, c).value, (int, float)) else None,
            }
            for stage in stage_keys:
                v = ws_cp.cell(dr + config.DATA_ROW_OFFSETS[stage], c).value
                entry[stage] = float(v) if isinstance(v, (int, float)) else None
            # Soil moisture: source switches by year
            if yr == 2026:
                ca_row = casma_row_by_monday.get(d)
                if ca_row is not None:
                    entry["top_vs"] = _safe_float(ws_ca.cell(ca_row, 3).value)
                    entry["top_s"]  = _safe_float(ws_ca.cell(ca_row, 4).value)
                    entry["top_a"]  = _safe_float(ws_ca.cell(ca_row, 5).value)
                    entry["top_su"] = _safe_float(ws_ca.cell(ca_row, 6).value)
                    entry["sub_vs"] = _safe_float(ws_ca.cell(ca_row, 7).value)
                    entry["sub_s"]  = _safe_float(ws_ca.cell(ca_row, 8).value)
                    entry["sub_a"]  = _safe_float(ws_ca.cell(ca_row, 9).value)
                    entry["sub_su"] = _safe_float(ws_ca.cell(ca_row, 10).value)
                else:
                    for k in ("top_vs", "top_s", "top_a", "top_su",
                              "sub_vs", "sub_s", "sub_a", "sub_su"):
                        entry[k] = None
            else:
                for key, off_key in [
                    ("top_vs", "topsoil_vs"), ("top_s", "topsoil_s"),
                    ("top_a", "topsoil_a"), ("top_su", "topsoil_su"),
                    ("sub_vs", "subsoil_vs"), ("sub_s", "subsoil_s"),
                    ("sub_a", "subsoil_a"), ("sub_su", "subsoil_su"),
                ]:
                    v = ws_cp.cell(dr + config.DATA_ROW_OFFSETS[off_key], c).value
                    entry[key] = _safe_float(v)
            weeks_out.append(entry)
        history_weekly[str(yr)] = weeks_out

    # ---- Assemble + write ----
    snapshot = {
        "as_of": date.today().isoformat(),
        "district": "EC Iowa (NASS District 60)",
        "counties": [n for fips, (n, _) in config.EC_IOWA_COUNTIES.items()],
        "total_corn_acres": config.TOTAL_CORN_ACRES,
        "forecast": forecast,
        "history": history,
        "history_weekly": history_weekly,
        "conditions": conditions,
        "model": {
            "yield": dict(ym),
            "casma_to_nass": config.CASMA_NASS_SUBSTRESS_CALIBRATION,
        },
    }

    OUT.write_text(json.dumps(snapshot, indent=2, default=str))
    print(f"wrote {OUT}  ({OUT.stat().st_size:,} bytes)")
    print(f"  as_of: {snapshot['as_of']}")
    print(f"  history entries: {len(history)}")
    print(f"  indicative 2026 forecast: {forecast['indicative_bu_ac']} bu/ac "
          f"[{forecast['band_low_95']}, {forecast['band_high_95']}]")
    if latest_casma:
        print(f"  latest CASMA: wk{latest_casma['iso_week']} ({latest_casma['monday']})")


if __name__ == "__main__":
    main()
